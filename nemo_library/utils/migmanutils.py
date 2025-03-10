import importlib
import json
import logging
import os
import re
import openpyxl
import pandas as pd
import configparser

from nemo_library.utils.config import Config
from nemo_library.features.projects import getImportedColumns, getProjectList
from nemo_library.utils.utils import get_internal_name


def initializeFolderStructure(
    project_path: str,
) -> None:

    folders = [
        "templates",
        "mappings",
        "srcdata",
        "other",
        "to_proalpha",
        "to_customer",
    ]
    for folder in folders:
        os.makedirs(os.path.join(project_path, folder), exist_ok=True)


def getMappingFilePath(projectname: str, local_project_path: str) -> str:
    return os.path.join(local_project_path, "mappings", f"{projectname}.csv")


def load_database() -> pd.DataFrame:
    with importlib.resources.open_binary(
        "nemo_library.templates", "migmantemplates.pkl"
    ) as file:
        df = pd.read_pickle(file)

    return df


def getProjectName(project: str, addon: str, postfix: str) -> str:
    return f"{project}{" " + addon if addon else ""}{(" (" + postfix + ")") if postfix else ""}"


def getNEMOStepsFrompAMigrationStatusFile(file: str) -> list[str]:
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook["Status Datenübernahme"]

    data = []
    for row in worksheet.iter_rows(
        min_row=10, max_row=300, min_col=1, max_col=10, values_only=True
    ):
        data.append(row)

    # Create a DataFrame from the extracted data
    columns = [
        worksheet.cell(row=9, column=i).value for i in range(1, 11)
    ]  # Headers in row 9
    dataframe = pd.DataFrame(data, columns=columns)

    # Drop rows where "Importreihenfolge" is NaN or empty
    if "Importreihenfolge" in dataframe.columns:
        dataframe = dataframe.dropna(subset=["Importreihenfolge"])
    else:
        raise ValueError("The column 'Importreihenfolge' does not exist in the data.")

    if "Name des Importprograms / Name der Erfassungsmaske" in dataframe.columns:
        nemosteps = dataframe[dataframe["Migrationsart"] == "NEMO"][
            "Name des Importprograms / Name der Erfassungsmaske"
        ].to_list()

        nemosteps = [x.title().strip() for x in nemosteps]
        replacements = {
            "European Article Numbers": "Global Trade Item Numbers",
            "Part-Storage Areas Relationship": "Part-Storage Areas Relationships",
            "Sales Tax Id": "Sales Tax ID",
            "Mrp Parameters": "MRP Parameters",
            "Sales Units Of Measure": "Sales Units of Measure",
            "Standard Boms (Header Data)": "Standard BOMs (Header Data)",
            "Standard Boms (Line Data)": "Standard BOMs (Line Data)",
            "Routings (Standard Boms)": "Routings (Standard BOMs)",
            "Bills Of Materials For Operations (Routings Production)": "Bills of Materials for Operations (Routings Production)",
        }

        nemosteps = [
            replacements[item] if item in replacements else item for item in nemosteps
        ]

        return nemosteps
    else:
        raise ValueError(
            "The column 'Name des Importprograms / Name der Erfassungsmaske' does not exist in the data."
        )


def getMappingRelations(config: Config) -> pd.DataFrame:

    # get configuration
    mapping_fields = config.get_migman_mapping_fields()
    additional_fields = config.get_migman_additional_fields()
    synonym_fields = config.get_migman_synonym_fields()

    # get data projects
    projectList = getProjectList(config=config)["displayName"].to_list()
    projectList = [
        project
        for project in projectList
        if not project.startswith("Mapping")
        and not project in ("Business Processes", "Master Data")
    ]

    # scan projects for fields
    data = []
    for project in projectList:

        logging.info(f"scan project '{project}' for mapping fields...")

        # get list of fields
        imported_columns_df = getImportedColumns(config=config, projectname=project)

        # remove (xxx)
        def remove_brackets_if_present(name):
            pattern = r"\(\d{3}\)$"
            if re.search(pattern, name):
                return re.sub(pattern, "", name).strip()
            return name

        imported_columns_df["displayName_cleaned"] = imported_columns_df[
            "displayName"
        ].apply(remove_brackets_if_present)
        imported_columns = imported_columns_df["displayName_cleaned"].to_list()

        # let's search the fields now
        for field in mapping_fields:
            # Check if the mapping field or any of its synonyms exists in imported_columns
            matching_field = None
            if field in imported_columns:
                matching_field = field
            else:
                for synonym in synonym_fields.get(field, []):
                    if synonym in imported_columns:
                        matching_field = synonym
                        break

            # If the mapping field or one of its synonyms is found
            if matching_field:

                # Check if all additional fields are also present
                additional_fields_present = all(
                    additional_field in imported_columns
                    for additional_field in additional_fields.get(field, [])
                )
                if additional_fields_present:

                    # we have cut of the (...) for easier handling. Now we have add them back again and add useful information for further processing
                    matching_field_display_name = imported_columns_df[
                        imported_columns_df["displayName_cleaned"] == matching_field
                    ]["displayName"].iloc[0]
                    matching_field_internal_name = imported_columns_df[
                        imported_columns_df["displayName_cleaned"] == matching_field
                    ]["internalName"].iloc[0]

                    additional_field_information = []
                    for additional_field in additional_fields.get(field, []):
                        additional_field_display_name = imported_columns_df[
                            imported_columns_df["displayName_cleaned"]
                            == additional_field
                        ]["displayName"].iloc[0]
                        additional_field_internal_name = imported_columns_df[
                            imported_columns_df["displayName_cleaned"]
                            == additional_field
                        ]["internalName"].iloc[0]
                        additional_field_information.append(
                            (
                                additional_field_display_name,
                                additional_field_internal_name,
                            )
                        )

                    # Save the data for this mapping field
                    data.append(
                        {
                            "project": project,
                            "mapping_field": field,
                            "matching_field_display_name": matching_field_display_name,
                            "matching_field_internal_name": matching_field_internal_name,
                            "additional_fields": additional_field_information,
                        }
                    )

    logging.info(f"mapping related fields found: {json.dumps(data,indent=2)}")
    return pd.DataFrame(data)


def sqlQueryInMappingTable(
    config: Config,
    field: str,
    newProject: bool,
    mappingrelationsdf: pd.DataFrame,
) -> str:

    projects = mappingrelationsdf["project"].to_list()
    display_names = mappingrelationsdf["matching_field_display_name"].to_list()
    internal_names = mappingrelationsdf["matching_field_internal_name"].to_list()
    additional_fields = mappingrelationsdf["additional_fields"].to_list()
    additional_fields_defined = config.get_migman_additional_fields()
    additional_field_global_definition = additional_fields_defined.get(field, [])

    # setup CTEs to load data from source projects
    ctes = []
    for project, display_name, internal_name, additional_fields in zip(
        projects, display_names, internal_names, additional_fields
    ):

        subselect = [f'{internal_name} AS "source {field}"']
        if any(additional_fields):
            for (
                additional_field_label,
                additional_field_internal_name,
            ), additional_field_definition in zip(
                additional_fields, additional_field_global_definition
            ):
                subselect.extend(
                    [
                        f'{additional_field_internal_name} AS "source {additional_field_definition}"'
                    ]
                )

        ctes.append(
            f"""CTE_{get_internal_name(project)} AS (
    SELECT DISTINCT
        {"\n\t,".join(subselect)}
    FROM 
        $schema.PROJECT_{get_internal_name(project)}
)"""
        )

    # global CTE to UNION ALL everything
    source_fields = [f'"source {field}"']
    for additional_field in additional_field_global_definition:
        source_fields.append(f'"source {additional_field}"')

    cteallfrags = [
        f"""SELECT
        {"\n\t, ".join(source_fields)} from CTE_{get_internal_name(project)} """
        for project in projects
    ]

    joinfrags = [
        f'mapping.{get_internal_name(field.strip('"'))} = cte.{field}'
        for field in source_fields
    ]

    # build the final query
    query = f""" WITH {"\n, ".join(ctes)}
, CTEALL AS (
    {"\nUNION ALL\n\t".join(cteallfrags)}
)
, CTEALLDISTINCT AS (
    SELECT DISTINCT
        {"\n\t, ".join(source_fields)}
    FROM   
        CTEALL  
)
SELECT
    cte.{"\n\t, cte.".join(source_fields)}
    , {"NULL" if newProject else f"mapping.TARGET_{get_internal_name(field)}"} AS "target {field}"
FROM    
    CTEALLDISTINCT cte
"""
    if not newProject:
        query += f"""LEFT JOIN
    $schema.$table mapping
ON  
    {"\n\tAND ".join(joinfrags)}"""

    return query
